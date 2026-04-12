from __future__ import annotations

import io
import json
import re
from collections import defaultdict
from datetime import datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils.cell import range_boundaries

TEMPLATE_ROWS = 19
TEMPLATE_COLS = 20
_TEMPLATE_TITLE_CELL = (2, 2)
_TITLE_MERGE_RANGE = "B2:T2"
_SCREEN_TOP_CELL = (5, 6)
_SCREEN_BOTTOM_CELL = (19, 6)
_SCREEN_TOP_MERGE_RANGE = "F5:Q5"
_SCREEN_BOTTOM_MERGE_RANGE = "F19:Q19"
_GOOGLE_MIRROR_START_ROW = 5
_GOOGLE_MIRROR_START_COL = 2
_GOOGLE_MIRROR_NUM_ROWS = 14
_GOOGLE_MIRROR_NUM_COLS = 19

# 座席表（テンプレ）の座標に合わせた 13テーブル x 6席 の配置定義
# スクリーン視点で千鳥配置: 前4 / 中5 / 後4
# テーブル番号は以下の順（左→右）
# 1:前1, 2:中1, 3:後1, 4:前2, 5:中2, 6:後2, ... , 13:中5
_FRONT_ROWS = (8, 9, 10)
_MIDDLE_ROWS = (12, 13, 14)
_BACK_ROWS = (16, 17, 18)
_FRONT_COL_PAIRS = ((4, 6), (8, 10), (12, 14), (16, 18))
_MIDDLE_COL_PAIRS = ((2, 4), (6, 8), (10, 12), (14, 16), (18, 20))
_BACK_COL_PAIRS = ((4, 6), (8, 10), (12, 14), (16, 18))
_TABLE_POSITION_ORDER: tuple[tuple[str, int], ...] = (
    ("front", 0),
    ("middle", 0),
    ("back", 0),
    ("front", 1),
    ("middle", 1),
    ("back", 1),
    ("front", 2),
    ("middle", 2),
    ("back", 2),
    ("front", 3),
    ("middle", 3),
    ("back", 3),
    ("middle", 4),
)

_GOOGLE_SCOPES = ("https://www.googleapis.com/auth/spreadsheets",)
_EXCEL_INVALID_TITLE = re.compile(r"[\\/*?:\[\]]")
_GOOGLE_INVALID_TITLE = re.compile(r"[\\/?*\[\]:]")
_FILE_INVALID = re.compile(r"[\\/:*?\"<>|]")
_DEFAULT_TEMPLATE_SHEET_NAME = "座席表（テンプレ）"
_SPECIAL_LABEL_KEYWORDS = ("スクリーン", "ホワイトボード", "メイン", "アシスタント")
_NORMAL_VIEW_SUFFIX = "（受講生視点）"
_MIRRORED_VIEW_SUFFIX = "（講師視点）"
_LEGACY_MIRROR_SUFFIX = "（反転）"

_TITLE_FONT = Font(name="Meiryo UI", size=14, bold=True)
_SCREEN_FONT = Font(name="Meiryo UI", size=13, bold=True)
_SEAT_FONT = Font(name="Meiryo UI", size=10, bold=True)
_CENTER = Alignment(horizontal="center", vertical="center")
_CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)
_THIN_SIDE = Side(style="thin", color="000000")
_THIN_BORDER = Border(left=_THIN_SIDE, right=_THIN_SIDE, top=_THIN_SIDE, bottom=_THIN_SIDE)
_TITLE_FILL = PatternFill(fill_type="solid", fgColor="FFF200")
_SCREEN_FILL = PatternFill(fill_type="solid", fgColor="F7F7F7")
_SEAT_FILL = PatternFill(fill_type="solid", fgColor="E9E9E9")


def default_daily_sheet_name(reference_dt: datetime | None = None) -> str:
    dt = reference_dt or datetime.now()
    return f"{dt.month}月{dt.day}日座席表"


def normalize_base_sheet_name(value: str) -> str:
    text = value.strip()
    base = text if text else default_daily_sheet_name()
    for suffix in (_NORMAL_VIEW_SUFFIX, _MIRRORED_VIEW_SUFFIX, _LEGACY_MIRROR_SUFFIX):
        if base.endswith(suffix):
            base = base[: -len(suffix)].strip()
    return base or default_daily_sheet_name()


def default_table_layout_rows(table_count: int = 13) -> list[list[int]]:
    front = [no for no in (1, 4, 7, 10) if no <= table_count]
    middle = [no for no in (2, 5, 8, 11, 13) if no <= table_count]
    back = [no for no in (3, 6, 9, 12) if no <= table_count]
    return [front, middle, back]


def build_layout_payload(
    rows: list[dict[str, Any]],
    base_sheet_name: str,
    include_skill: bool = True,
    table_layout_rows: list[list[int]] | None = None,
) -> dict[str, Any]:
    base_name = normalize_base_sheet_name(base_sheet_name)
    normal_name = f"{base_name}{_NORMAL_VIEW_SUFFIX}"
    mirrored_name = f"{base_name}{_MIRRORED_VIEW_SUFFIX}"
    return {
        "normal_name": normal_name,
        "mirrored_name": mirrored_name,
        "normal_grid": build_template_layout_grid(
            rows,
            normal_name,
            mirrored=False,
            include_skill=include_skill,
            table_layout_rows=table_layout_rows,
        ),
        "mirrored_grid": build_template_layout_grid(
            rows,
            mirrored_name,
            mirrored=True,
            include_skill=include_skill,
            table_layout_rows=table_layout_rows,
        ),
    }


def safe_download_filename(value: str, suffix: str = "") -> str:
    cleaned = _FILE_INVALID.sub("_", value.strip())
    cleaned = cleaned.strip(" .")
    base = cleaned or "seating_layout"
    return f"{base}{suffix}" if suffix else base


def build_template_layout_grid(
    rows: list[dict[str, Any]],
    sheet_title: str,
    mirrored: bool,
    include_skill: bool = True,
    table_layout_rows: list[list[int]] | None = None,
) -> list[list[str]]:
    grid = [["" for _ in range(TEMPLATE_COLS)] for _ in range(TEMPLATE_ROWS)]
    grid[_TEMPLATE_TITLE_CELL[0] - 1][_TEMPLATE_TITLE_CELL[1] - 1] = sheet_title
    screen_row, screen_col = _SCREEN_BOTTOM_CELL if mirrored else _SCREEN_TOP_CELL
    grid[screen_row - 1][screen_col - 1] = "スクリーン"

    table_slots = _resolve_table_slots(table_layout_rows)
    members_by_table: dict[int, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        try:
            table_no = int(row["table_no"])
        except (TypeError, ValueError, KeyError):
            continue
        if table_no in table_slots:
            members_by_table[table_no].append(row)

    for table_no in sorted(table_slots.keys()):
        slots = table_slots[table_no]
        members = sorted(members_by_table.get(table_no, []), key=_member_sort_key)
        # mirrored=True は「前から後ろへ見る視点」（左右反転ではなく前後方向を反転）
        active_slots = [_front_to_back_slot(slot) for slot in slots] if mirrored else slots

        for idx, (r_idx, c_idx) in enumerate(active_slots):
            text = (
                _format_member_cell(members[idx], include_skill=include_skill)
                if idx < len(members)
                else ""
            )
            grid[r_idx - 1][c_idx - 1] = text

    return grid


def export_layouts_to_excel_bytes(
    normal_grid: list[list[str]],
    mirrored_grid: list[list[str]],
    normal_sheet_name: str,
    mirrored_sheet_name: str,
) -> bytes:
    workbook = Workbook()
    used_titles: set[str] = set()

    normal_title = _unique_title(
        _sanitize_title(normal_sheet_name, max_len=31, invalid_re=_EXCEL_INVALID_TITLE),
        used_titles,
        max_len=31,
    )
    ws_normal = workbook.active
    ws_normal.title = normal_title
    _write_grid_to_worksheet(ws_normal, normal_grid)
    _style_excel_worksheet(ws_normal, mirrored=False)

    mirrored_title = _unique_title(
        _sanitize_title(mirrored_sheet_name, max_len=31, invalid_re=_EXCEL_INVALID_TITLE),
        used_titles,
        max_len=31,
    )
    ws_mirrored = workbook.create_sheet(title=mirrored_title)
    _write_grid_to_worksheet(ws_mirrored, mirrored_grid)
    _style_excel_worksheet(ws_mirrored, mirrored=True)

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def publish_layouts_to_google_sheets(
    service_account_json_bytes: bytes,
    spreadsheet_ref: str,
    template_sheet_name: str,
    normal_sheet_name: str,
    normal_grid: list[list[str]],
    mirrored_sheet_name: str,
    mirrored_grid: list[list[str]],
) -> dict[str, str]:
    import gspread
    from google.oauth2.service_account import Credentials

    if not service_account_json_bytes:
        raise ValueError("サービスアカウントJSONが空です。")
    _ = mirrored_grid

    try:
        account_info = json.loads(service_account_json_bytes.decode("utf-8-sig"))
    except json.JSONDecodeError as exc:
        raise ValueError("サービスアカウントJSONの読み取りに失敗しました。") from exc

    spreadsheet_id = extract_spreadsheet_id(spreadsheet_ref)
    if not spreadsheet_id:
        raise ValueError("出力先スプレッドシートID(URL)を指定してください。")

    sheet_name = template_sheet_name.strip()
    if not sheet_name:
        raise ValueError("テンプレートシート名を指定してください。")

    credentials = Credentials.from_service_account_info(account_info, scopes=list(_GOOGLE_SCOPES))
    client = gspread.authorize(credentials)
    spreadsheet = client.open_by_key(spreadsheet_id)
    available_titles = [ws.title for ws in spreadsheet.worksheets()]
    resolved_sheet_name = _resolve_template_sheet_name(sheet_name, available_titles)
    if not resolved_sheet_name:
        preview = ", ".join(available_titles[:10]) if available_titles else "（シートなし）"
        raise ValueError(
            f"テンプレートシート「{sheet_name}」が見つかりません。"
            f"存在するシート名: {preview}"
        )
    template_ws = spreadsheet.worksheet(resolved_sheet_name)

    existing_titles = {ws.title for ws in spreadsheet.worksheets()}
    normal_title = _unique_title(
        _sanitize_title(normal_sheet_name, max_len=100, invalid_re=_GOOGLE_INVALID_TITLE),
        existing_titles,
        max_len=100,
    )
    mirrored_title = _unique_title(
        _sanitize_title(mirrored_sheet_name, max_len=100, invalid_re=_GOOGLE_INVALID_TITLE),
        existing_titles,
        max_len=100,
    )

    ws_normal = template_ws.duplicate(new_sheet_name=normal_title)

    probe_cols = max(TEMPLATE_COLS, 40)
    template_range = f"A1:{_to_col_label(probe_cols)}{TEMPLATE_ROWS}"
    template_raw = template_ws.get(template_range)
    used_cols = max((len(row) for row in template_raw), default=TEMPLATE_COLS)
    write_cols = max(TEMPLATE_COLS, used_cols)
    template_matrix = _normalize_matrix(template_raw, rows=TEMPLATE_ROWS, cols=write_cols)

    kana_lookup = _extract_name_kana_lookup(template_matrix)
    normal_grid_with_kana = _apply_name_kana_fallback(normal_grid, kana_lookup)

    normal_matrix = [row[:] for row in template_matrix]
    _overlay_non_empty_grid(
        normal_matrix,
        normal_grid_with_kana,
        limit_cols=TEMPLATE_COLS,
        skip_special_labels=True,
    )
    _overlay_seat_cells_exact(
        base_matrix=normal_matrix,
        grid=normal_grid_with_kana,
    )

    update_range = f"A1:{_to_col_label(write_cols)}{TEMPLATE_ROWS}"
    ws_normal.update(range_name=update_range, values=normal_matrix)
    ws_mirrored = ws_normal.duplicate(new_sheet_name=mirrored_title)
    _apply_google_region_180_mirror(
        spreadsheet=spreadsheet,
        source_ws=ws_normal,
        target_ws=ws_mirrored,
        start_row=_GOOGLE_MIRROR_START_ROW,
        start_col=_GOOGLE_MIRROR_START_COL,
        num_rows=_GOOGLE_MIRROR_NUM_ROWS,
        num_cols=_GOOGLE_MIRROR_NUM_COLS,
    )
    ws_mirrored.update_acell(
        f"{_to_col_label(_TEMPLATE_TITLE_CELL[1])}{_TEMPLATE_TITLE_CELL[0]}",
        mirrored_title,
    )

    return {
        "spreadsheet_id": spreadsheet_id,
        "requested_template_sheet_name": sheet_name,
        "used_template_sheet_name": resolved_sheet_name,
        "normal_sheet_name": normal_title,
        "normal_sheet_url": f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/edit#gid={ws_normal.id}",
        "mirrored_sheet_name": mirrored_title,
        "mirrored_sheet_url": f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/edit#gid={ws_mirrored.id}",
    }


def extract_spreadsheet_id(value: str) -> str:
    text = value.strip()
    if not text:
        return ""
    match = re.search(r"/spreadsheets/d/([a-zA-Z0-9-_]+)", text)
    if match:
        return match.group(1)
    return text


def _build_table_slots() -> dict[int, list[tuple[int, int]]]:
    slots: dict[int, list[tuple[int, int]]] = {}
    zone_layout: dict[str, tuple[tuple[int, ...], tuple[tuple[int, int], ...]]] = {
        "front": (_FRONT_ROWS, _FRONT_COL_PAIRS),
        "middle": (_MIDDLE_ROWS, _MIDDLE_COL_PAIRS),
        "back": (_BACK_ROWS, _BACK_COL_PAIRS),
    }
    for table_no, (zone, idx) in enumerate(_TABLE_POSITION_ORDER, start=1):
        if zone not in zone_layout:
            raise ValueError(f"unknown zone: {zone}")
        row_group, col_pairs = zone_layout[zone]
        left_col, right_col = col_pairs[idx]
        table_slots: list[tuple[int, int]] = []
        for row in row_group:
            table_slots.append((row, left_col))
            table_slots.append((row, right_col))
        slots[table_no] = table_slots
    return slots


def _build_reverse_map(values: list[int]) -> dict[int, int]:
    return {value: values[-(idx + 1)] for idx, value in enumerate(values)}


_TABLE_SLOTS = _build_table_slots()
_SEAT_COORDS = {(row, col) for slots in _TABLE_SLOTS.values() for row, col in slots}
_SEAT_ROWS = sorted({row for row, _ in _SEAT_COORDS})
_SEAT_COLS = sorted({col for _, col in _SEAT_COORDS})
_FRONT_ROW_MAP = _build_reverse_map(_SEAT_ROWS)
_FRONT_COL_MAP = _build_reverse_map(_SEAT_COLS)


def _resolve_table_slots(table_layout_rows: list[list[int]] | None) -> dict[int, list[tuple[int, int]]]:
    if not table_layout_rows:
        return dict(_TABLE_SLOTS)

    custom = _build_table_slots_from_layout_rows(table_layout_rows)
    return custom if custom else dict(_TABLE_SLOTS)


def _build_table_slots_from_layout_rows(
    table_layout_rows: list[list[int]],
) -> dict[int, list[tuple[int, int]]]:
    row_groups: tuple[tuple[int, ...], ...] = (_FRONT_ROWS, _MIDDLE_ROWS, _BACK_ROWS)
    row_col_pairs: tuple[tuple[tuple[int, int], ...], ...] = (
        _FRONT_COL_PAIRS,
        _MIDDLE_COL_PAIRS,
        _BACK_COL_PAIRS,
    )

    slots: dict[int, list[tuple[int, int]]] = {}
    for row_idx in range(min(3, len(table_layout_rows))):
        table_numbers = table_layout_rows[row_idx]
        col_pairs = row_col_pairs[row_idx]
        row_group = row_groups[row_idx]
        for pos_idx, table_no_raw in enumerate(table_numbers):
            if pos_idx >= len(col_pairs):
                break
            try:
                table_no = int(table_no_raw)
            except (TypeError, ValueError):
                continue
            if table_no <= 0 or table_no in slots:
                continue
            left_col, right_col = col_pairs[pos_idx]
            table_slots: list[tuple[int, int]] = []
            for row in row_group:
                table_slots.append((row, left_col))
                table_slots.append((row, right_col))
            slots[table_no] = table_slots
    return slots


def _member_sort_key(row: dict[str, Any]) -> tuple[int, str]:
    try:
        sid = int(row.get("student_id", 0))
    except (TypeError, ValueError):
        sid = 0
    return sid, str(row.get("name", ""))


def _format_member_cell(row: dict[str, Any], include_skill: bool) -> str:
    name_kana = str(row.get("name_kana", "")).strip()
    name = str(row.get("name", "")).strip()
    company = str(row.get("company", "")).strip()
    skill = str(row.get("skill_level", "")).strip()
    parts = [part for part in (name_kana, name, company) if part]
    if include_skill and skill:
        parts.append(skill)
    return "\n".join(parts)


def _front_to_back_slot(slot: tuple[int, int]) -> tuple[int, int]:
    row, col = slot
    mapped_row = _FRONT_ROW_MAP.get(row, row)
    mapped_col = _FRONT_COL_MAP.get(col, col)
    return mapped_row, mapped_col


def _write_grid_to_worksheet(worksheet: Any, grid: list[list[str]]) -> None:
    for row_idx, row_values in enumerate(grid, start=1):
        for col_idx, value in enumerate(row_values, start=1):
            worksheet.cell(row=row_idx, column=col_idx, value=value)


def _style_excel_worksheet(worksheet: Any, mirrored: bool) -> None:
    _set_excel_dimensions(worksheet)
    worksheet.merge_cells(_TITLE_MERGE_RANGE)
    screen_merge_range = _SCREEN_BOTTOM_MERGE_RANGE if mirrored else _SCREEN_TOP_MERGE_RANGE
    worksheet.merge_cells(screen_merge_range)

    _style_range(
        worksheet,
        _TITLE_MERGE_RANGE,
        fill=_TITLE_FILL,
        border=_THIN_BORDER,
        alignment=_CENTER,
        font=_TITLE_FONT,
    )
    _style_range(
        worksheet,
        screen_merge_range,
        fill=_SCREEN_FILL,
        border=_THIN_BORDER,
        alignment=_CENTER,
        font=_SCREEN_FONT,
    )

    for row, col in _SEAT_COORDS:
        cell = worksheet.cell(row=row, column=col)
        cell.fill = _SEAT_FILL
        cell.border = _THIN_BORDER
        cell.alignment = _CENTER_WRAP
        cell.font = _SEAT_FONT


def _set_excel_dimensions(worksheet: Any) -> None:
    for col in range(1, TEMPLATE_COLS + 1):
        label = _to_col_label(col)
        worksheet.column_dimensions[label].width = 17.0 if col in _SEAT_COLS else 2.8

    for row in range(1, TEMPLATE_ROWS + 1):
        worksheet.row_dimensions[row].height = 20.0
    for row in _SEAT_ROWS:
        worksheet.row_dimensions[row].height = 50.0

    worksheet.row_dimensions[2].height = 28.0
    worksheet.row_dimensions[5].height = 24.0
    worksheet.row_dimensions[19].height = 24.0


def _style_range(
    worksheet: Any,
    cell_range: str,
    fill: PatternFill,
    border: Border,
    alignment: Alignment,
    font: Font,
) -> None:
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.fill = fill
            cell.border = border
            cell.alignment = alignment
            cell.font = font


def _sanitize_title(value: str, max_len: int, invalid_re: re.Pattern[str]) -> str:
    cleaned = invalid_re.sub(" ", value.strip())
    cleaned = re.sub(r"\s+", " ", cleaned).strip()
    if not cleaned:
        cleaned = "座席表"
    return cleaned[:max_len]


def _unique_title(base: str, used: set[str], max_len: int) -> str:
    candidate = base[:max_len]
    if candidate not in used:
        used.add(candidate)
        return candidate

    counter = 2
    while True:
        suffix = f" ({counter})"
        trimmed = base[: max_len - len(suffix)]
        candidate = f"{trimmed}{suffix}"
        if candidate not in used:
            used.add(candidate)
            return candidate
        counter += 1


def _to_col_label(col_num: int) -> str:
    label = ""
    n = col_num
    while n > 0:
        n, rem = divmod(n - 1, 26)
        label = chr(65 + rem) + label
    return label


def _normalize_matrix(raw: list[list[Any]], rows: int, cols: int) -> list[list[str]]:
    matrix = [["" for _ in range(cols)] for _ in range(rows)]
    for r_idx in range(min(rows, len(raw))):
        row = raw[r_idx]
        for c_idx in range(min(cols, len(row))):
            value = row[c_idx]
            matrix[r_idx][c_idx] = "" if value is None else str(value)
    return matrix


def _overlay_non_empty_grid(
    base_matrix: list[list[str]],
    grid: list[list[str]],
    limit_cols: int,
    skip_special_labels: bool = False,
) -> None:
    max_rows = min(len(base_matrix), len(grid))
    max_cols = min(limit_cols, len(base_matrix[0]) if base_matrix else 0)
    for r_idx in range(max_rows):
        row_values = grid[r_idx]
        for c_idx in range(min(max_cols, len(row_values))):
            value = str(row_values[c_idx]) if row_values[c_idx] is not None else ""
            if value:
                if skip_special_labels:
                    normalized = _normalize_label_text(value)
                    if any(keyword in normalized for keyword in _SPECIAL_LABEL_KEYWORDS):
                        continue
                base_matrix[r_idx][c_idx] = value


def _overlay_seat_cells_exact(base_matrix: list[list[str]], grid: list[list[str]]) -> None:
    if not base_matrix or not grid:
        return
    for row_no, col_no in _SEAT_COORDS:
        r_idx = row_no - 1
        c_idx = col_no - 1
        if r_idx < 0 or c_idx < 0:
            continue
        if r_idx >= len(base_matrix) or r_idx >= len(grid):
            continue
        if c_idx >= len(base_matrix[r_idx]) or c_idx >= len(grid[r_idx]):
            continue
        value = grid[r_idx][c_idx]
        base_matrix[r_idx][c_idx] = "" if value is None else str(value)


def _rotate_matrix_region_180(
    matrix: list[list[str]],
    start_row: int,
    start_col: int,
    num_rows: int,
    num_cols: int,
) -> None:
    if not matrix or num_rows <= 0 or num_cols <= 0:
        return
    row_start = max(0, start_row - 1)
    col_start = max(0, start_col - 1)
    row_end = min(len(matrix), row_start + num_rows)
    if row_start >= row_end:
        return
    col_end = min(min(len(row) for row in matrix), col_start + num_cols)
    if col_start >= col_end:
        return

    region = [row[col_start:col_end] for row in matrix[row_start:row_end]]
    reversed_region = [list(reversed(row)) for row in reversed(region)]
    for r_idx, row_values in enumerate(reversed_region, start=row_start):
        matrix[r_idx][col_start:col_end] = row_values


def _apply_google_region_180_mirror(
    spreadsheet: Any,
    source_ws: Any,
    target_ws: Any,
    start_row: int,
    start_col: int,
    num_rows: int,
    num_cols: int,
) -> None:
    mirror_range = _a1_range(start_row, start_col, num_rows, num_cols)

    source_values_raw = source_ws.get(mirror_range)
    source_values = _normalize_matrix(source_values_raw, rows=num_rows, cols=num_cols)
    reversed_values = [list(reversed(row)) for row in reversed(source_values)]

    target_ws.unmerge_cells(mirror_range)
    target_ws.batch_clear([mirror_range])
    target_ws.update(range_name=mirror_range, values=reversed_values)

    format_requests = _build_google_format_rotate_requests(
        source_sheet_id=int(source_ws.id),
        target_sheet_id=int(target_ws.id),
        start_row=start_row,
        start_col=start_col,
        num_rows=num_rows,
        num_cols=num_cols,
    )
    if format_requests:
        spreadsheet.batch_update({"requests": format_requests})

    _reverse_google_dimensions(
        spreadsheet=spreadsheet,
        source_ws=source_ws,
        target_ws=target_ws,
        start_row=start_row,
        start_col=start_col,
        num_rows=num_rows,
        num_cols=num_cols,
    )

    source_merges = _get_google_sheet_merges(
        spreadsheet=spreadsheet,
        sheet_id=int(source_ws.id),
    )
    merged_entries = _filter_merges_in_region(
        source_merges=source_merges,
        start_row=start_row,
        start_col=start_col,
        num_rows=num_rows,
        num_cols=num_cols,
    )

    merge_format_requests: list[dict[str, Any]] = []
    for merge in merged_entries:
        dst_merge = _rotate_merge_180(
            merge=merge,
            start_row=start_row,
            start_col=start_col,
            num_rows=num_rows,
            num_cols=num_cols,
        )
        target_ws.merge_cells(_grid_range_to_a1(dst_merge))
        merge_format_requests.append(
            {
                "copyPaste": {
                    "source": {
                        "sheetId": int(source_ws.id),
                        "startRowIndex": int(merge["startRowIndex"]),
                        "endRowIndex": int(merge["endRowIndex"]),
                        "startColumnIndex": int(merge["startColumnIndex"]),
                        "endColumnIndex": int(merge["endColumnIndex"]),
                    },
                    "destination": {
                        "sheetId": int(target_ws.id),
                        "startRowIndex": int(dst_merge["startRowIndex"]),
                        "endRowIndex": int(dst_merge["endRowIndex"]),
                        "startColumnIndex": int(dst_merge["startColumnIndex"]),
                        "endColumnIndex": int(dst_merge["endColumnIndex"]),
                    },
                    "pasteType": "PASTE_FORMAT",
                    "pasteOrientation": "NORMAL",
                }
            }
        )

        src_value = _get_single_display_value(source_ws, _grid_range_to_a1(merge))
        if src_value:
            target_ws.update_acell(
                _cell_a1(dst_merge["startRowIndex"] + 1, dst_merge["startColumnIndex"] + 1),
                src_value,
            )
    if merge_format_requests:
        spreadsheet.batch_update({"requests": merge_format_requests})


def _a1_range(start_row: int, start_col: int, num_rows: int, num_cols: int) -> str:
    end_row = start_row + num_rows - 1
    end_col = start_col + num_cols - 1
    return f"{_to_col_label(start_col)}{start_row}:{_to_col_label(end_col)}{end_row}"


def _cell_a1(row: int, col: int) -> str:
    return f"{_to_col_label(col)}{row}"


def _grid_range_to_a1(grid_range: dict[str, int]) -> str:
    start_row = int(grid_range["startRowIndex"]) + 1
    end_row = int(grid_range["endRowIndex"])
    start_col = int(grid_range["startColumnIndex"]) + 1
    end_col = int(grid_range["endColumnIndex"])
    return f"{_to_col_label(start_col)}{start_row}:{_to_col_label(end_col)}{end_row}"


def _build_google_format_rotate_requests(
    source_sheet_id: int,
    target_sheet_id: int,
    start_row: int,
    start_col: int,
    num_rows: int,
    num_cols: int,
) -> list[dict[str, Any]]:
    row_start0 = start_row - 1
    col_start0 = start_col - 1
    requests: list[dict[str, Any]] = []
    for r_idx in range(num_rows):
        for c_idx in range(num_cols):
            src_row0 = row_start0 + r_idx
            src_col0 = col_start0 + c_idx
            dst_row0 = row_start0 + (num_rows - 1 - r_idx)
            dst_col0 = col_start0 + (num_cols - 1 - c_idx)
            requests.append(
                {
                    "copyPaste": {
                        "source": {
                            "sheetId": int(source_sheet_id),
                            "startRowIndex": src_row0,
                            "endRowIndex": src_row0 + 1,
                            "startColumnIndex": src_col0,
                            "endColumnIndex": src_col0 + 1,
                        },
                        "destination": {
                            "sheetId": int(target_sheet_id),
                            "startRowIndex": dst_row0,
                            "endRowIndex": dst_row0 + 1,
                            "startColumnIndex": dst_col0,
                            "endColumnIndex": dst_col0 + 1,
                        },
                        "pasteType": "PASTE_FORMAT",
                        "pasteOrientation": "NORMAL",
                    }
                }
            )
    return requests


def _reverse_google_dimensions(
    spreadsheet: Any,
    source_ws: Any,
    target_ws: Any,
    start_row: int,
    start_col: int,
    num_rows: int,
    num_cols: int,
) -> None:
    source_title = str(source_ws.title).replace("'", "''")
    mirror_range = _a1_range(start_row, start_col, num_rows, num_cols)
    try:
        metadata = spreadsheet.fetch_sheet_metadata(
            params={
                "includeGridData": True,
                "ranges": [f"'{source_title}'!{mirror_range}"],
            }
        )
    except Exception:
        return

    source_sheet_data: dict[str, Any] | None = None
    for sheet in metadata.get("sheets", []):
        properties = sheet.get("properties", {})
        if int(properties.get("sheetId", -1)) == int(source_ws.id):
            source_sheet_data = sheet
            break
    if source_sheet_data is None:
        return

    data_blocks = source_sheet_data.get("data", [])
    if not data_blocks:
        return
    block = data_blocks[0]
    col_meta = block.get("columnMetadata", []) or []
    row_meta = block.get("rowMetadata", []) or []

    col_sizes = [int(meta.get("pixelSize")) for meta in col_meta if meta.get("pixelSize")]
    row_sizes = [int(meta.get("pixelSize")) for meta in row_meta if meta.get("pixelSize")]
    if not col_sizes and not row_sizes:
        return

    col_start0 = start_col - 1
    row_start0 = start_row - 1
    requests: list[dict[str, Any]] = []
    for idx, size in enumerate(col_sizes[:num_cols]):
        dst_col0 = col_start0 + (num_cols - 1 - idx)
        requests.append(
            {
                "updateDimensionProperties": {
                    "range": {
                        "sheetId": int(target_ws.id),
                        "dimension": "COLUMNS",
                        "startIndex": dst_col0,
                        "endIndex": dst_col0 + 1,
                    },
                    "properties": {"pixelSize": size},
                    "fields": "pixelSize",
                }
            }
        )

    for idx, size in enumerate(row_sizes[:num_rows]):
        dst_row0 = row_start0 + (num_rows - 1 - idx)
        requests.append(
            {
                "updateDimensionProperties": {
                    "range": {
                        "sheetId": int(target_ws.id),
                        "dimension": "ROWS",
                        "startIndex": dst_row0,
                        "endIndex": dst_row0 + 1,
                    },
                    "properties": {"pixelSize": size},
                    "fields": "pixelSize",
                }
            }
        )

    if requests:
        spreadsheet.batch_update({"requests": requests})


def _get_google_sheet_merges(spreadsheet: Any, sheet_id: int) -> list[dict[str, int]]:
    metadata = spreadsheet.fetch_sheet_metadata()
    for sheet in metadata.get("sheets", []):
        properties = sheet.get("properties", {})
        if int(properties.get("sheetId", -1)) == int(sheet_id):
            merges = sheet.get("merges", [])
            return [dict(merge) for merge in merges]
    return []


def _filter_merges_in_region(
    source_merges: list[dict[str, int]],
    start_row: int,
    start_col: int,
    num_rows: int,
    num_cols: int,
) -> list[dict[str, int]]:
    row_start = start_row - 1
    row_end = row_start + num_rows
    col_start = start_col - 1
    col_end = col_start + num_cols

    selected: list[dict[str, int]] = []
    for merge in source_merges:
        m_row_start = int(merge.get("startRowIndex", -1))
        m_row_end = int(merge.get("endRowIndex", -1))
        m_col_start = int(merge.get("startColumnIndex", -1))
        m_col_end = int(merge.get("endColumnIndex", -1))
        if (
            m_row_start >= row_start
            and m_row_end <= row_end
            and m_col_start >= col_start
            and m_col_end <= col_end
        ):
            selected.append(
                {
                    "startRowIndex": m_row_start,
                    "endRowIndex": m_row_end,
                    "startColumnIndex": m_col_start,
                    "endColumnIndex": m_col_end,
                }
            )
    return selected


def _rotate_merge_180(
    merge: dict[str, int],
    start_row: int,
    start_col: int,
    num_rows: int,
    num_cols: int,
) -> dict[str, int]:
    row_start = start_row - 1
    col_start = start_col - 1
    m_row_start = int(merge["startRowIndex"])
    m_row_end = int(merge["endRowIndex"])
    m_col_start = int(merge["startColumnIndex"])
    m_col_end = int(merge["endColumnIndex"])
    m_rows = m_row_end - m_row_start
    m_cols = m_col_end - m_col_start

    rel_row = m_row_start - row_start
    rel_col = m_col_start - col_start

    dst_row_start = row_start + (num_rows - rel_row - m_rows)
    dst_col_start = col_start + (num_cols - rel_col - m_cols)
    return {
        "startRowIndex": dst_row_start,
        "endRowIndex": dst_row_start + m_rows,
        "startColumnIndex": dst_col_start,
        "endColumnIndex": dst_col_start + m_cols,
    }


def _get_single_display_value(worksheet: Any, a1_range: str) -> str:
    values = worksheet.get(a1_range)
    if not values:
        return ""
    first_row = values[0]
    if not first_row:
        return ""
    return str(first_row[0] or "")


def _mirror_special_labels(matrix: list[list[str]], max_cols: int) -> None:
    if not matrix:
        return
    rows = len(matrix)
    labels: list[tuple[int, int, str]] = []
    for r_idx in range(rows):
        for c_idx in range(min(max_cols, len(matrix[r_idx]))):
            text = str(matrix[r_idx][c_idx]).strip()
            if not text:
                continue
            normalized = _normalize_label_text(text)
            if any(keyword in normalized for keyword in _SPECIAL_LABEL_KEYWORDS):
                labels.append((r_idx + 1, c_idx + 1, text))

    for r, c, _ in labels:
        matrix[r - 1][c - 1] = ""

    seat_col_min = min(_SEAT_COLS)
    seat_col_max = max(_SEAT_COLS)
    label_rows = [row for row, _, _ in labels]
    label_cols = [col for _, col, _ in labels]
    mirror_row_min = min(_SCREEN_TOP_CELL[0], min(label_rows))
    mirror_row_max = max(_SCREEN_BOTTOM_CELL[0], max(label_rows))
    mirror_col_min = min(seat_col_min, min(label_cols))
    mirror_col_max = max(seat_col_max, max(label_cols))

    for row, col, text in labels:
        mapped_row, mapped_col = _map_position_180(
            row=row,
            col=col,
            row_min=mirror_row_min,
            row_max=mirror_row_max,
            col_min=mirror_col_min,
            col_max=mirror_col_max,
        )
        if 1 <= mapped_row <= rows and 1 <= mapped_col <= max_cols:
            matrix[mapped_row - 1][mapped_col - 1] = text


def _map_position_180(
    row: int,
    col: int,
    row_min: int,
    row_max: int,
    col_min: int,
    col_max: int,
) -> tuple[int, int]:
    mapped_row = row_min + row_max - row
    mapped_col = col_min + col_max - col
    return mapped_row, mapped_col


def _normalize_label_text(value: str) -> str:
    return re.sub(r"\s+", "", value).strip()


def _is_kana_text(value: str) -> bool:
    text = _normalize_label_text(value)
    if not text:
        return False
    allowed = set("ぁあぃいぅうぇえぉおかがきぎくぐけげこごさざしじすずせぜそぞただちぢっつづてでとどなにぬねのはばぱひびぴふぶぷへべぺほぼぽまみむめもゃやゅゆょよらりるれろゎわゐゑをんーァアィイゥウェエォオカガキギクグケゲコゴサザシジスズセゼソゾタダチヂッツヅテデトドナニヌネノハバパヒビピフブプヘベペホボポマミムメモャヤュユョヨラリルレロヮワヰヱヲンヴ・ ")
    return all(ch in allowed for ch in text)


def _extract_name_kana_lookup(template_matrix: list[list[str]]) -> dict[tuple[str, str], str]:
    lookup: dict[tuple[str, str], str] = {}
    for row in template_matrix:
        for cell in row:
            text = str(cell or "").strip()
            if not text or "\n" not in text:
                continue
            lines = [line.strip() for line in text.split("\n") if line.strip()]
            if len(lines) < 2:
                continue

            kana = ""
            name = ""
            company = ""
            if len(lines) >= 3 and _is_kana_text(lines[0]):
                kana = lines[0]
                name = lines[1]
                company = lines[2]
            elif len(lines) >= 2 and _is_kana_text(lines[0]):
                kana = lines[0]
                name = lines[1]
                company = lines[2] if len(lines) >= 3 else ""

            if kana and name:
                name_key = _normalize_person_key(name)
                company_key = _normalize_person_key(company)
                lookup[(name_key, company_key)] = kana
                if (name_key, "") not in lookup:
                    lookup[(name_key, "")] = kana
    return lookup


def _apply_name_kana_fallback(
    grid: list[list[str]], kana_lookup: dict[tuple[str, str], str]
) -> list[list[str]]:
    if not kana_lookup:
        return [row[:] for row in grid]

    updated = [row[:] for row in grid]
    for row_no, col_no in _SEAT_COORDS:
        r_idx = row_no - 1
        c_idx = col_no - 1
        if r_idx < 0 or r_idx >= len(updated):
            continue
        if c_idx < 0 or c_idx >= len(updated[r_idx]):
            continue
        text = str(updated[r_idx][c_idx] or "").strip()
        if not text:
            continue

        lines = [line.strip() for line in text.split("\n") if line.strip()]
        if len(lines) < 2:
            continue
        if _is_kana_text(lines[0]):
            continue

        name = _normalize_person_key(lines[0])
        company = _normalize_person_key(lines[1] if len(lines) >= 2 else "")
        kana = kana_lookup.get((name, company)) or kana_lookup.get((name, ""))
        if not kana:
            continue

        rebuilt = [kana, lines[0]]
        rebuilt.extend(lines[1:])
        updated[r_idx][c_idx] = "\n".join(rebuilt)
    return updated


def _normalize_person_key(value: str) -> str:
    return " ".join(str(value).replace("　", " ").strip().split())


def _resolve_template_sheet_name(requested_name: str, available_titles: list[str]) -> str:
    if requested_name in available_titles:
        return requested_name

    normalized_to_original = {_normalize_sheet_title(title): title for title in available_titles}
    normalized_requested = _normalize_sheet_title(requested_name)
    if normalized_requested in normalized_to_original:
        return normalized_to_original[normalized_requested]

    trimmed_date_prefix = re.sub(r"^\d{1,2}月\d{1,2}日", "", requested_name).strip()
    if trimmed_date_prefix in available_titles:
        return trimmed_date_prefix
    normalized_trimmed = _normalize_sheet_title(trimmed_date_prefix)
    if normalized_trimmed in normalized_to_original:
        return normalized_to_original[normalized_trimmed]

    if _DEFAULT_TEMPLATE_SHEET_NAME in available_titles:
        return _DEFAULT_TEMPLATE_SHEET_NAME

    template_like = [title for title in available_titles if "テンプレ" in title]
    if len(template_like) == 1:
        return template_like[0]

    return ""


def _normalize_sheet_title(value: str) -> str:
    return value.replace(" ", "").replace("　", "").strip()
